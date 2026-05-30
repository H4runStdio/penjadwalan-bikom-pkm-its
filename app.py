import io
import random
import re
from datetime import datetime, timedelta

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ─────────────────────────────────────────────
#  KONSTANTA
# ─────────────────────────────────────────────
KOLOM_WAJIB_TIM   = ["nrp", "nama ketua", "bidang pkm"]
KOLOM_WAJIB_DOSEN = ["nama lengkap", "bidang pkm"]

CONTOH_TIM = pd.DataFrame(
    {
        "NRP":        ["5026221001", "5026221002"],
        "Nama Ketua": ["Budi Santoso", "Ani Rahayu"],
        "Bidang PKM": ["PKM-RE", "PKM-KC"],
    }
)

CONTOH_DOSEN = pd.DataFrame(
    {
        "Nama Lengkap": ["Dr. Andi", "Dr. Budi"],
        "Bidang PKM":   ["PKM-RE, PKM-KC", "PKM-T"],
        "Lokasi":       ["Ruang A101", "Ruang B202"],
        "Senin":        ["08.00 - 10.00, 13.00 - 15.00", "09.00 - 11.00"],
        "Selasa":       ["Tidak Bersedia", "08.00 - 10.00"],
    }
)


# ─────────────────────────────────────────────
#  CLEANING & VALIDASI
# ─────────────────────────────────────────────
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [c.strip().lower() for c in df.columns]
    return df


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(how="all").reset_index(drop=True)
    for col in df.select_dtypes(include="object").columns:
        df[col] = (
            df[col]
            .astype(str)
            .str.strip()
            .str.replace("–", "-", regex=False)
            .str.replace("—", "-", regex=False)
            .str.replace("−", "-", regex=False)
            .str.replace(r"\s+", " ", regex=True)
        )
    return df


def validasi_kolom(df: pd.DataFrame, kolom_wajib: list[str], label: str) -> list[str]:
    ada    = set(df.columns.str.strip().str.lower())
    kurang = [k for k in kolom_wajib if k not in ada]
    if kurang:
        return [
            f"File **{label}** kekurangan kolom: "
            + ", ".join(f"`{k.title()}`" for k in kurang)
        ]
    return []


def validasi_isi_tim(df: pd.DataFrame) -> list[str]:
    errors = []
    for col, label in [("nrp", "NRP"), ("nama ketua", "Nama Ketua"), ("bidang pkm", "Bidang PKM")]:
        if df[col].isnull().any() or (df[col] == "nan").any():
            errors.append(f"Ada baris di file tim dengan **{label} kosong**.")
    return errors


def resolve_kolom_hari(df_columns: list, kolom_hari: dict) -> dict:
    """
    Cocokkan nama hari ke kolom aktual di DataFrame (partial match, case-insensitive).
    Dipanggil SATU KALI di luar loop, bukan per-baris.
    """
    cols_lower = {c.lower(): c for c in df_columns}
    resolved   = {}
    for nama_hari, kol_input in kolom_hari.items():
        kol_lower = kol_input.lower()
        if kol_lower in cols_lower:
            resolved[nama_hari] = cols_lower[kol_lower]
        else:
            matches = [c_orig for c_low, c_orig in cols_lower.items() if kol_lower in c_low]
            resolved[nama_hari] = matches[0] if matches else None
    return resolved


def _is_kosong(val: str) -> bool:
    return val.strip().lower() in ("", "nan", "none", "tidak bersedia", "-", "n/a", "na")


def validasi_isi_dosen(df: pd.DataFrame, kolom_hari: dict) -> list[str]:
    errors   = []
    resolved = resolve_kolom_hari(list(df.columns), kolom_hari)

    tidak_ketemu = [h for h, kol in resolved.items() if kol is None]
    if tidak_ketemu:
        errors.append(
            "File dosen tidak memiliki kolom untuk hari: "
            + ", ".join(f"**{h}**" for h in tidak_ketemu)
            + ". Pastikan nama kolom di sidebar sesuai."
        )

    pola_jam = re.compile(r"^\d{2}\.\d{2} - \d{2}\.\d{2}(, \d{2}\.\d{2} - \d{2}\.\d{2})*$")
    for _, row in df.iterrows():
        for nama_hari, kol_aktual in resolved.items():
            if kol_aktual is None:
                continue
            val = str(row.get(kol_aktual, "")).strip()
            if _is_kosong(val):
                continue
            if not pola_jam.match(val):
                errors.append(
                    f"Format jam tidak valid untuk dosen **{row.get('nama lengkap', '?')}** "
                    f"hari **{nama_hari}**: `{val}` - "
                    "gunakan format `HH.MM - HH.MM` (pisah koma jika lebih dari satu sesi)."
                )
    return errors


def cek_bidang_mismatch(df_tim: pd.DataFrame, df_dosen: pd.DataFrame) -> list[str]:
    semua_bidang_dosen: set[str] = set()
    for val in df_dosen["bidang pkm"]:
        for b in str(val).split(","):
            semua_bidang_dosen.add(b.strip().lower())

    warnings = []
    for bidang in df_tim["bidang pkm"].unique():
        if str(bidang).lower() not in semua_bidang_dosen:
            jumlah = (df_tim["bidang pkm"] == bidang).sum()
            warnings.append(
                f"Bidang **{bidang}** ({jumlah} tim) tidak memiliki dosen pembimbing yang sesuai, "
                "tim ini dipastikan tidak terplot."
            )
    return warnings


def deduplikasi_tim(df_tim: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """
    Hapus baris dengan NRP + Bidang PKM yang PERSIS sama (duplikasi sejati).
    NRP sama tapi Bidang PKM berbeda tetap dipertahankan (diplot keduanya).
    Kembalikan (df_bersih, jumlah_dihapus).
    """
    sebelum = len(df_tim)
    df_clean = df_tim.drop_duplicates(subset=["nrp", "bidang pkm"], keep="first").reset_index(drop=True)
    return df_clean, sebelum - len(df_clean)


# ─────────────────────────────────────────────
#  PARSING JAM
# ─────────────────────────────────────────────
def buat_slot(jam_mulai: str, jam_selesai: str, durasi: int) -> list[str]:
    fmt   = "%H.%M"
    start = datetime.strptime(jam_mulai, fmt)
    end   = datetime.strptime(jam_selesai, fmt)
    slots = []
    while start + timedelta(minutes=durasi) <= end:
        nxt = start + timedelta(minutes=durasi)
        slots.append(f"{start.strftime('%H:%M')}-{nxt.strftime('%H:%M')}")
        start = nxt
    return slots


def parse_sesi(jam_text: str, durasi: int) -> list[dict]:
    if pd.isna(jam_text) or _is_kosong(str(jam_text)):
        return []
    sesi_list = []
    for bagian in str(jam_text).split(","):
        bagian = bagian.strip()
        if " - " not in bagian:
            continue
        mulai, selesai = [x.strip() for x in bagian.split(" - ", 1)]
        try:
            slots = buat_slot(mulai, selesai, durasi)
            if slots:
                sesi_list.append({"range": f"{mulai} - {selesai}", "slots": slots})
        except ValueError:
            pass
    return sesi_list


def jam_mulai_dt(sesi_range: str) -> datetime:
    return datetime.strptime(sesi_range.split(" - ")[0].strip(), "%H.%M")


def hitung_sisa_jam(sesi_range: str, slot_terpakai: int, semua_slots: list[str]) -> str | None:
    """Kembalikan string sisa jam, atau None jika semua sudah terpakai."""
    total = len(semua_slots)
    if slot_terpakai >= total:
        return None
    # Sisa: dari jam akhir slot terakhir yang terpakai sampai akhir sesi
    _, selesai_range = sesi_range.split(" - ", 1)
    if slot_terpakai == 0:
        mulai_range = sesi_range.split(" - ")[0].strip()
        return f"{mulai_range} - {selesai_range.strip()}"
    last_slot_end = semua_slots[slot_terpakai - 1].split("-")[1]  # format HH.MM
    return f"{last_slot_end} - {selesai_range.strip()}"


# ─────────────────────────────────────────────
#  BUILD SLOT DOSEN
# ─────────────────────────────────────────────
def build_slot_dosen(df_dosen: pd.DataFrame, kolom_hari: dict, durasi: int) -> dict:
    """
    Bangun struktur slot dosen.
    resolve_kolom_hari dipanggil SATU KALI di luar loop baris.
    """
    resolved   = resolve_kolom_hari(list(df_dosen.columns), kolom_hari)
    slot_dosen = {}

    for _, row in df_dosen.iterrows():
        nama         = row["nama lengkap"]
        bidang_list  = [b.strip() for b in str(row["bidang pkm"]).split(",")]
        lokasi_raw   = row.get("lokasi", "")
        lokasi       = "" if (pd.isna(lokasi_raw) or str(lokasi_raw).strip().lower() in ("nan", "none", "")) else str(lokasi_raw).strip()

        slot_dosen[nama] = {
            "bidang":         [b.lower() for b in bidang_list],
            "bidang_display": bidang_list,
            "lokasi":         lokasi,
            "jadwal":         {},
        }

        for nama_hari, kol in resolved.items():
            if kol is None or kol not in row.index:
                continue
            sesi_list = parse_sesi(row[kol], durasi)
            if sesi_list:
                slot_dosen[nama]["jadwal"][nama_hari] = [
                    {
                        "range":       s["range"],
                        "slots":       s["slots"],
                        "slot_index":  0,
                    }
                    for s in sesi_list
                ]
    return slot_dosen


def hitung_sisa_kapasitas(slot_dosen: dict, maks: int) -> dict[str, int]:
    result = {}
    for nama, data in slot_dosen.items():
        total = 0
        for sesi_list in data["jadwal"].values():
            for sesi in sesi_list:
                total += max(0, min(maks, len(sesi["slots"])) - sesi["slot_index"])
        result[nama] = total
    return result


# ─────────────────────────────────────────────
#  ALGORITMA PENJADWALAN
# ─────────────────────────────────────────────
def hitung_pilihan_dosen_per_bidang(slot_dosen: dict) -> dict[str, int]:
    """
    Hitung berapa banyak dosen yang tersedia untuk setiap bidang PKM.
    Digunakan untuk sorting least-options-first.
    """
    bidang_count: dict[str, int] = {}
    for data in slot_dosen.values():
        for b in data["bidang"]:
            bidang_count[b] = bidang_count.get(b, 0) + 1
    return bidang_count


def run_scheduling(
    df_tim:        pd.DataFrame,
    slot_dosen:    dict,
    maks_per_sesi: int,
    random_seed:   int,
) -> list[dict]:
    """
    Greedy dengan dua level optimasi:

    1. Least-options-first (LOF): bidang yang punya sedikit dosen diproses lebih dulu
       agar tidak kalah berebut slot dengan bidang yang punya banyak dosen.
    2. Dalam satu bidang, urutan FIFO (urutan baris di file) tetap dipertahankan
       sehingga pendaftar pertama tetap prioritas mendapat slot.
    3. Pemilihan dosen: acak + sort by sisa kapasitas (load balancing).
    """
    rng = random.Random(random_seed)

    # Hitung jumlah dosen per bidang — tetap dihitung sekali di awal
    pilihan_per_bidang = hitung_pilihan_dosen_per_bidang(slot_dosen)

    # Kelompokkan tim per bidang, pertahankan urutan asli (FIFO) dalam tiap bidang
    # dengan menyimpan index baris asli sebagai tiebreaker
    from collections import defaultdict
    tim_per_bidang: dict[str, list[tuple[int, object]]] = defaultdict(list)
    for idx, (_, tim) in enumerate(df_tim.iterrows()):
        bidang = str(tim["bidang pkm"]).strip().lower()
        tim_per_bidang[bidang].append((idx, tim))

    # Tentukan dosen eksklusif vs dosen multi-bidang
    # Dosen eksklusif: hanya melayani 1 bidang — bidangnya tidak bersaing
    # Dosen multi-bidang: melayani >1 bidang — menjadi rebutan
    dosen_per_bidang: dict[str, list[str]] = {}
    for nama_dosen, data in slot_dosen.items():
        for b in data["bidang"]:
            dosen_per_bidang.setdefault(b, []).append(nama_dosen)

    def punya_dosen_eksklusif(bidang: str) -> bool:
        """True jika bidang ini punya setidaknya satu dosen yang HANYA melayani bidang ini."""
        for nama_dosen in dosen_per_bidang.get(bidang, []):
            if len(slot_dosen[nama_dosen]["bidang"]) == 1:
                return True
        return False

    # Pisah bidang menjadi 3 grup berdasarkan tingkat kekritisan:
    # Grup 1 (paling kritis): hanya punya 1-2 dosen AND tidak punya dosen eksklusif
    #         → rebutan penuh, harus diproses paling awal
    # Grup 2: punya dosen eksklusif → aman, tapi tetap sebelum bidang besar
    # Grup 3: punya banyak dosen (>=3) → diproses FIFO normal
    THRESHOLD_LANGKA = 3

    def grup_bidang(bidang: str) -> int:
        n = pilihan_per_bidang.get(bidang, 0)
        if n < THRESHOLD_LANGKA and not punya_dosen_eksklusif(bidang):
            return 0   # kritis, rebutan
        if punya_dosen_eksklusif(bidang):
            return 1   # aman eksklusif
        return 2       # banyak pilihan

    urutan_bidang = sorted(
        tim_per_bidang.keys(),
        key=lambda b: (grup_bidang(b), pilihan_per_bidang.get(b, 0), len(tim_per_bidang[b]))
    )

    # Susun antrian:
    # - Grup 0 (kritis) diproses penuh terlebih dahulu
    # - Grup 1 dan 2 diproses FIFO berdasarkan index baris asli
    antrian_kritis: list[tuple[int, object]] = []
    antrian_normal: list[tuple[int, object]] = []

    for bidang in urutan_bidang:
        if grup_bidang(bidang) == 0:
            antrian_kritis.extend(tim_per_bidang[bidang])
        else:
            antrian_normal.extend(tim_per_bidang[bidang])

    # Antrian normal diurutkan kembali by index asli (FIFO global untuk non-kritis)
    antrian_normal.sort(key=lambda x: x[0])

    antrian: list[tuple[int, object]] = antrian_kritis + antrian_normal

    # Proses antrian
    hasil: list[dict] = []
    for _idx, tim in antrian:
        bidang_tim = str(tim["bidang pkm"]).strip().lower()
        assigned   = False

        kandidat = [
            nama for nama, data in slot_dosen.items()
            if bidang_tim in data["bidang"]
        ]
        rng.shuffle(kandidat)
        sisa = hitung_sisa_kapasitas(slot_dosen, maks_per_sesi)
        kandidat.sort(key=lambda n: sisa[n], reverse=True)

        for nama_dosen in kandidat:
            if assigned:
                break
            data = slot_dosen[nama_dosen]
            for nama_hari, sesi_list in data["jadwal"].items():
                if assigned:
                    break
                for sesi in sesi_list:
                    if sesi["slot_index"] >= maks_per_sesi:
                        continue
                    if sesi["slot_index"] >= len(sesi["slots"]):
                        continue
                    jam = sesi["slots"][sesi["slot_index"]]
                    sesi["slot_index"] += 1
                    hasil.append(
                        {
                            "Dosen":  nama_dosen,
                            "Hari":   nama_hari,
                            "Sesi":   sesi["range"],
                            "Jam":    jam,
                            "Ketua":  tim["nama ketua"],
                            "NRP":    str(tim["nrp"]),
                            "Bidang": tim["bidang pkm"],
                            "Lokasi": data["lokasi"],
                        }
                    )
                    assigned = True
                    break

    return hasil


# ─────────────────────────────────────────────
#  HELPER STYLE EXCEL
# ─────────────────────────────────────────────
def _make_border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


_BORDER  = None   # lazy-init per wb session
_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _cell(ws, r: int, c: int, value=None, fill=None, font=None, align=None):
    """Tulis nilai + style ke satu sel. Semua parameter opsional."""
    cell = ws.cell(row=r, column=c)
    if value is not None:
        cell.value = value
    cell.border    = _make_border()
    cell.alignment = align or _CENTER
    cell.font      = font if font is not None else FONT_NORMAL
    if fill:
        cell.fill = fill
    return cell


def _merge(ws, r1: int, c1: int, r2: int, c2: int,
           value=None, fill=None, font=None, align=None):
    """
    Merge cells dan terapkan border ke SEMUA sel dalam range,
    bukan hanya sel kiri-atas — ini yang menyebabkan border tidak sempurna.
    """
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    b = _make_border()
    f = font if font is not None else FONT_NORMAL
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = b
            if fill:
                cell.fill = fill
            cell.font = f
    top_left = ws.cell(row=r1, column=c1)
    if value is not None:
        top_left.value = value
    top_left.alignment = align or _CENTER


def _auto_col_width(ws, max_width: int = 40):
    from openpyxl.cell.cell import MergedCell
    for col_cells in ws.columns:
        # Skip kolom yang diawali MergedCell (tidak punya column_letter)
        first = next((c for c in col_cells if not isinstance(c, MergedCell)), None)
        if first is None:
            continue
        length = max(
            (len(str(cell.value or "")) for cell in col_cells if not isinstance(cell, MergedCell)),
            default=8,
        )
        ws.column_dimensions[first.column_letter].width = min(length + 4, max_width)


# ─────────────────────────────────────────────
#  EXPORT EXCEL — sheet-level functions
# ─────────────────────────────────────────────
FILL_DOSEN  = _make_fill("3D85C6")
FILL_HARI   = _make_fill("9FC5E8")
FILL_LOKASI = _make_fill("FFF2CC")
FILL_HEADER = _make_fill("D9D9D9")
FILL_BELUM  = _make_fill("FCE5CD")

FONT_WHITE_BOLD = Font(name="Arial", bold=True, color="FFFFFF")
FONT_BOLD       = Font(name="Arial", bold=True)
FONT_NORMAL     = Font(name="Arial")


def _sheet_jadwal_hari(wb: Workbook, hari: str, daftar: list, maks_per_sesi: int,
                       max_table_per_row: int):
    """Buat satu sheet untuk satu hari bimbingan."""
    TABLE_W = 5          # No | Jam | Nama Ketua | Bidang  (+1 kolom gap antar tabel)
    HDR_H   = 4          # baris dosen + hari/sesi + lokasi + header kolom
    TABLE_H = HDR_H + maks_per_sesi

    ws = wb.create_sheet(hari)
    ws.sheet_view.showGridLines = True

    daftar_sorted = sorted(daftar, key=lambda x: jam_mulai_dt(x[0]))

    for table_idx, (sesi, dosen, items) in enumerate(daftar_sorted):
        sc = 1 + (table_idx % max_table_per_row) * TABLE_W
        sr = 1 + (table_idx // max_table_per_row) * (TABLE_H + 2)

        # Baris 1 — nama dosen (span 4 kolom: No, Jam, Nama Ketua, Bidang)
        _merge(ws, sr, sc, sr, sc + 3,
               value=dosen, fill=FILL_DOSEN, font=FONT_WHITE_BOLD)

        # Baris 2 — hari + sesi (format: Rabu, 09:00 - 13:00)
        sesi_display = sesi.replace(".", ":")
        _merge(ws, sr + 1, sc, sr + 1, sc + 3,
               value=f"{hari}, {sesi_display}", fill=FILL_HARI)

        # Baris 3 — lokasi (kosong jika tidak ada, tanpa em dash)
        lokasi_val = (items[0]["Lokasi"] if items else "") or ""
        _merge(ws, sr + 2, sc, sr + 2, sc + 3,
               value=lokasi_val, fill=FILL_LOKASI)

        # Baris 4 — header kolom (4 kolom, tanpa NRP)
        for i, h in enumerate(["No", "Jam", "Nama Ketua", "Bidang"]):
            _cell(ws, sr + 3, sc + i, value=h, fill=FILL_HEADER, font=FONT_BOLD)

        # Baris data
        for i in range(maks_per_sesi):
            r = sr + 4 + i
            _cell(ws, r, sc, value=i + 1)
            if i < len(items):
                _cell(ws, r, sc + 1, value=items[i]["Jam"])
                _cell(ws, r, sc + 2, value=items[i]["Ketua"], align=_LEFT)
                _cell(ws, r, sc + 3, value=items[i]["Bidang"])
            else:
                for c in range(sc + 1, sc + 4):
                    _cell(ws, r, c)

    _auto_col_width(ws)
    # Kolom "No" setiap tabel — paksa sempit
    from openpyxl.utils import get_column_letter
    for table_idx in range(len(daftar_sorted)):
        no_col = 1 + (table_idx % max_table_per_row) * TABLE_W
        ws.column_dimensions[get_column_letter(no_col)].width = 5


def _sheet_tim_belum_terplot(wb: Workbook, df_tim: pd.DataFrame, terplot_nrp_bidang: set):
    """Sheet daftar tim yang tidak berhasil dijadwalkan."""
    ws = wb.create_sheet("Tim Belum Terplot")
    headers = ["No", "NRP", "Nama Ketua", "Bidang PKM"]
    for i, h in enumerate(headers, 1):
        _cell(ws, 1, i, value=h, fill=FILL_HEADER, font=FONT_BOLD)

    row_idx = 2
    nomor   = 1
    for _, r in df_tim.iterrows():
        key = (str(r["nrp"]), str(r["bidang pkm"]))
        if key not in terplot_nrp_bidang:
            _cell(ws, row_idx, 1, value=nomor)
            _cell(ws, row_idx, 2, value=str(r["nrp"]),         align=_LEFT)
            _cell(ws, row_idx, 3, value=str(r["nama ketua"]),  align=_LEFT)
            _cell(ws, row_idx, 4, value=str(r["bidang pkm"]),  align=_LEFT)
            row_idx += 1
            nomor   += 1

    _auto_col_width(ws)


def _sheet_dosen_belum_terplot(wb: Workbook, slot_dosen: dict):
    """Sheet sisa slot dosen yang belum terisi."""
    ws      = wb.create_sheet("Dosen Belum Terplot")
    headers = ["No", "Nama Dosen", "Hari", "Sesi Jam", "Sisa Jam Kosong", "Lokasi"]
    for i, h in enumerate(headers, 1):
        _cell(ws, 1, i, value=h, fill=FILL_HEADER, font=FONT_BOLD)

    row = 2
    no  = 1
    for nama_dosen, data in slot_dosen.items():
        for nama_hari, sesi_list in data["jadwal"].items():
            for sesi in sesi_list:
                sisa_str = hitung_sisa_jam(sesi["range"], sesi["slot_index"], sesi["slots"])
                if sisa_str:
                    vals = [no, nama_dosen, nama_hari, sesi["range"],
                            sisa_str, data["lokasi"] or ""]
                    for i, v in enumerate(vals, 1):
                        _cell(ws, row, i, value=v, align=_LEFT)
                    row += 1
                    no  += 1

    _auto_col_width(ws)


def _sheet_rekap(wb: Workbook, df_hasil: pd.DataFrame):
    """Sheet rekap statistik hasil penjadwalan."""
    ws = wb.create_sheet("Rekap")
    if df_hasil.empty:
        ws.cell(1, 1, "Tidak ada data hasil penjadwalan.")
        return

    def tulis_block(start_row: int, judul: str, rows: list[list]) -> int:
        ws.cell(row=start_row, column=1, value=judul).font = FONT_BOLD
        r = start_row + 1
        for data_row in rows:
            is_header = (r == start_row + 1)
            fill = FILL_HEADER if is_header else None
            for c_idx, val in enumerate(data_row, 1):
                _cell(ws, r, c_idx, value=val,
                      fill=fill,
                      font=FONT_BOLD if is_header else None,
                      align=_LEFT)
            r += 1
        return r + 1  # baris kosong pemisah

    r = 1

    # 1. Tim per hari per bidang
    pivot = pd.pivot_table(
        df_hasil, index="Hari", columns="Bidang",
        values="Ketua", aggfunc="count", fill_value=0,
    )
    header = ["Hari"] + list(pivot.columns)
    rows   = [header] + [[h] + list(v) for h, v in pivot.iterrows()]
    r = tulis_block(r, "Jumlah Tim per Hari per Bidang PKM", rows)

    # 2. Tim per hari
    rows = [["Hari", "Jumlah Tim"]] + [
        [h, int(c)] for h, c in df_hasil.groupby("Hari")["Ketua"].count().items()
    ]
    r = tulis_block(r, "Jumlah Tim per Hari", rows)

    # 3. Dosen per hari
    rows = [["Hari", "Jumlah Dosen"]] + [
        [h, int(c)] for h, c in df_hasil.groupby("Hari")["Dosen"].nunique().items()
    ]
    r = tulis_block(r, "Jumlah Dosen per Hari", rows)

    # 4. Tim per dosen
    rows = [["Nama Dosen", "Jumlah Tim"]] + [
        [d, int(c)] for d, c in df_hasil.groupby("Dosen")["Ketua"].count().items()
    ]
    tulis_block(r, "Jumlah Tim per Dosen", rows)

    _auto_col_width(ws, max_width=50)


# ─────────────────────────────────────────────
#  EXPORT EXCEL — orkestrator
# ─────────────────────────────────────────────
def export_excel(
    hasil:           list[dict],
    slot_dosen:      dict,
    df_tim:          pd.DataFrame,
    maks_per_sesi:   int,
    max_table_per_row: int,
) -> bytes:
    wb = Workbook()
    del wb["Sheet"]

    # Kelompokkan hasil: hari → dosen → sesi → [item]
    grouped: dict = {}
    for h in hasil:
        (grouped
         .setdefault(h["Hari"], {})
         .setdefault(h["Dosen"], {})
         .setdefault(h["Sesi"], [])
         .append(h))

    # Sheet per hari
    for hari in sorted(grouped):
        daftar = [
            (sesi, dosen, items)
            for dosen, sd in grouped[hari].items()
            for sesi, items in sd.items()
        ]
        _sheet_jadwal_hari(wb, hari, daftar, maks_per_sesi, max_table_per_row)

    # Sheet tim belum terplot
    terplot_set = {(h["NRP"], h["Bidang"]) for h in hasil}
    _sheet_tim_belum_terplot(wb, df_tim, terplot_set)

    # Sheet dosen belum terplot
    _sheet_dosen_belum_terplot(wb, slot_dosen)

    # Sheet rekap
    _sheet_rekap(wb, pd.DataFrame(hasil) if hasil else pd.DataFrame())

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
#  KOMPONEN UI
# ─────────────────────────────────────────────
def _icon(name: str, size: int = 16, color: str = "currentColor") -> str:
    """
    Render satu Lucide icon sebagai HTML inline menggunakan unpkg CDN.
    Dipakai bersama st.markdown(..., unsafe_allow_html=True).
    """
    return (
        f'<img src="https://unpkg.com/lucide-static@latest/icons/{name}.svg" '
        f'width="{size}" height="{size}" '
        f'style="vertical-align:-3px;margin-right:6px;'
        f'filter:invert(0);opacity:0.85;" />'
    )


def _header_with_icon(icon_name: str, text: str, tag: str = "h3") -> None:
    """Tampilkan subheader/header dengan Lucide icon di depannya."""
    st.markdown(
        f'<{tag} style="display:flex;align-items:center;gap:6px;margin-bottom:0">'
        f'{_icon(icon_name, 20)}<span>{text}</span></{tag}>',
        unsafe_allow_html=True,
    )


def render_panduan():
    with st.expander(":clipboard: Panduan Format File Excel", expanded=False):
        st.markdown(
            """
Unggah dua file Excel dengan format di bawah ini.
Nama kolom **tidak case-sensitive** (huruf besar/kecil tidak berpengaruh).
            """
        )
        tab1, tab2 = st.tabs(["File Tim", "File Dosen"])

        with tab1:
            st.markdown("**Kolom wajib:**")
            st.markdown(
                "- `NRP` : Nomor Registrasi Pokok mahasiswa ketua tim  \n"
                "- `Nama Ketua` : Nama lengkap ketua tim  \n"
                "- `Bidang PKM` : Bidang PKM tim (harus sama persis dengan yang ada di file dosen)"
            )
            st.markdown(
                "**Catatan duplikasi NRP:**  \n"
                "Jika satu NRP muncul lebih dari sekali dengan **Bidang PKM berbeda**, "
                "keduanya akan dijadwalkan masing-masing.  \n"
                "Jika NRP dan Bidang PKM **sama persis**, baris duplikat akan dihapus otomatis (keep first)."
            )
            st.markdown("**Contoh data:**")
            st.dataframe(CONTOH_TIM, hide_index=True, use_container_width=True)

        with tab2:
            st.markdown("**Kolom wajib:**")
            st.markdown(
                "- `Nama Lengkap` : Nama dosen  \n"
                "- `Bidang PKM` : Bidang yang bisa dibimbing, pisah koma jika lebih dari satu  \n"
                "- `Lokasi` : Ruang/lokasi *(opsional, kosongkan jika belum ditentukan)*  \n"
                "- **Kolom hari** : Sesuai yang diisi di sidebar. "
                "Cukup tuliskan sebagian nama kolom, misalnya `Senin` sudah cukup untuk "
                "mencocokkan `Jam Kesediaan pada Senin, 12 Januari 2026`."
            )
            st.markdown(
                "**Format jam:** `HH.MM - HH.MM`  \n"
                "Lebih dari satu sesi per hari pisahkan dengan koma: `08.00 - 10.00, 13.00 - 15.00`  \n"
                "Jika tidak bersedia, isi `Tidak Bersedia` atau kosongkan."
            )
            st.markdown("**Contoh data:**")
            st.dataframe(CONTOH_DOSEN, hide_index=True, use_container_width=True)


def render_preview_kapasitas(df_tim: pd.DataFrame, slot_dosen: dict, maks_per_sesi: int):
    with st.expander(":bar_chart: Preview Kapasitas Sebelum Proses", expanded=True):
        total_tim = len(df_tim)
        total_kapasitas = sum(
            min(maks_per_sesi, len(s["slots"]))
            for data in slot_dosen.values()
            for sesi_list in data["jadwal"].values()
            for s in sesi_list
        )

        col1, col2, col3 = st.columns(3)
        col1.metric("Jumlah Tim", total_tim)
        col2.metric("Total Kapasitas Slot", total_kapasitas)
        delta = total_kapasitas - total_tim
        col3.metric(
            "Selisih",
            delta,
            delta=f"{'+' if delta >= 0 else ''}{delta}",
            delta_color="normal" if delta >= 0 else "inverse",
        )

        if delta < 0:
            st.warning(
                f"Kapasitas dosen ({total_kapasitas}) lebih sedikit dari jumlah tim ({total_tim}). "
                f"Diprediksi **{abs(delta)} tim tidak akan terplot**."
            )
        else:
            st.success(f"Kapasitas cukup. Ada {delta} slot tersisa setelah semua tim terplot.")

        st.markdown("**Kapasitas per Bidang PKM:**")
        bidang_tim = df_tim["bidang pkm"].value_counts().reset_index()
        bidang_tim.columns = ["Bidang PKM", "Jumlah Tim"]

        kapasitas_bidang: dict[str, int] = {}
        for data in slot_dosen.values():
            for bidang in data["bidang_display"]:
                for sesi_list in data["jadwal"].values():
                    for s in sesi_list:
                        kapasitas_bidang[bidang] = (
                            kapasitas_bidang.get(bidang, 0) + min(maks_per_sesi, len(s["slots"]))
                        )

        bidang_tim["Kapasitas Dosen"] = bidang_tim["Bidang PKM"].map(
            lambda b: kapasitas_bidang.get(b, 0)
        )
        bidang_tim["Status"] = bidang_tim.apply(
            lambda r: "Cukup" if r["Kapasitas Dosen"] >= r["Jumlah Tim"] else "Kurang",
            axis=1,
        )
        st.dataframe(bidang_tim, hide_index=True, use_container_width=True)


# ─────────────────────────────────────────────
#  MAIN APP
# ─────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="Penjadwalan Bimbingan Komunal",
        page_icon=":calendar:",
        layout="wide",
    )

    # Inject Lucide CDN sekali di awal
    st.html(
        '<script src="https://unpkg.com/lucide@latest"></script>'
        '<script>document.addEventListener("DOMContentLoaded",()=>lucide.createIcons())</script>'
    )

    st.markdown(
        f'{_icon("calendar-days", 28)} **Sistem Penjadwalan Bimbingan Komunal**',
        unsafe_allow_html=True,
    )
    st.caption("Penjadwalan otomatis berbasis ketersediaan dosen dan bidang PKM.")

    # ── Sidebar ──
    st.sidebar.markdown(
        f'### {_icon("settings",16)} Parameter Penjadwalan',
        unsafe_allow_html=True,
    )

    DURASI          = st.sidebar.number_input("Durasi per Tim (menit)",             10, 120, 20)
    MAKS_PER_SESI   = st.sidebar.number_input("Maks Tim per Dosen per Sesi",          1,  30, 12)
    MAX_TABLE_ROW   = st.sidebar.number_input("Jumlah Tabel per Baris (Excel)",        1,  10,  6)
    OUTPUT_FILE     = st.sidebar.text_input("Nama File Output", value="jadwal_bikom.xlsx")
    RANDOM_SEED     = st.sidebar.number_input(
        "Random Seed",
        min_value=0, max_value=9999, value=42,
        help="Angka yang sama menghasilkan urutan dosen yang sama. Ubah jika ingin variasi.",
    )

    st.sidebar.markdown("---")
    st.sidebar.markdown(
        f'### {_icon("calendar",16)} Hari Bimbingan',
        unsafe_allow_html=True,
    )

    JUMLAH_HARI = st.sidebar.number_input("Jumlah Hari", 1, 14, 5)
    HARI_BIMBINGAN: dict[str, str] = {}

    for i in range(JUMLAH_HARI):
        c1, c2 = st.sidebar.columns(2)
        nama_hari = c1.text_input(f"Nama Hari {i+1}", key=f"hari_{i}", placeholder="Senin")
        nama_kol  = c2.text_input(f"Kolom Excel {i+1}", key=f"kol_{i}", placeholder="Senin")
        if nama_hari.strip() and nama_kol.strip():
            HARI_BIMBINGAN[nama_hari.strip()] = nama_kol.strip()

    # ── Panduan ──
    render_panduan()
    st.divider()

    # ── Upload ──
    _header_with_icon("upload-cloud", "Upload File Data")
    col_up1, col_up2 = st.columns(2)
    with col_up1:
        file_tim   = st.file_uploader("File Data Tim (.xlsx)",   type=["xlsx"])
    with col_up2:
        file_dosen = st.file_uploader("File Data Dosen (.xlsx)", type=["xlsx"])

    if not (file_tim and file_dosen):
        st.info("Unggah kedua file untuk melanjutkan.")
        return

    # ── Baca & bersihkan ──
    try:
        df_tim_raw   = pd.read_excel(file_tim)
        df_dosen_raw = pd.read_excel(file_dosen)
    except Exception as e:
        st.error(f"Gagal membaca file Excel: {e}")
        return

    df_tim   = clean_dataframe(normalize_columns(df_tim_raw.copy()))
    df_dosen = clean_dataframe(normalize_columns(df_dosen_raw.copy()))

    # ── Validasi kolom ──
    errors: list[str] = []
    errors += validasi_kolom(df_tim,   KOLOM_WAJIB_TIM,   "Tim")
    errors += validasi_kolom(df_dosen, KOLOM_WAJIB_DOSEN, "Dosen")

    if errors:
        st.error("Struktur file tidak sesuai. Perbaiki terlebih dahulu:")
        for e in errors:
            st.markdown(f"- {e}")
        return

    # ── De-duplikasi tim (NRP + Bidang PKM sama persis) ──
    df_tim, n_dup = deduplikasi_tim(df_tim)
    if n_dup > 0:
        st.info(
            f"Ditemukan **{n_dup} baris duplikat** (NRP + Bidang PKM sama persis) dan telah dihapus. "
            "Tim dengan NRP sama tapi Bidang PKM berbeda tetap dijadwalkan masing-masing."
        )

    # ── Validasi hari ──
    if not HARI_BIMBINGAN:
        st.warning("Belum ada hari bimbingan yang dikonfigurasi di sidebar.")
        return

    kolom_hari_lower = {h: k.strip() for h, k in HARI_BIMBINGAN.items()}

    # ── Validasi isi ──
    warnings_isi: list[str] = []
    warnings_isi += validasi_isi_tim(df_tim)
    warnings_isi += validasi_isi_dosen(df_dosen, kolom_hari_lower)
    warnings_isi += cek_bidang_mismatch(df_tim, df_dosen)

    if warnings_isi:
        with st.expander(":warning: Peringatan Data (klik untuk lihat)", expanded=True):
            for w in warnings_isi:
                st.warning(w)
            st.markdown(
                "_Anda tetap bisa memproses, namun tim/dosen yang bermasalah mungkin tidak terplot._"
            )

    # ── Preview data ──
    _header_with_icon("table", "Preview Data")
    with st.expander("Lihat data yang terbaca", expanded=False):
        t1, t2 = st.tabs(["Data Tim", "Data Dosen"])
        with t1:
            st.dataframe(df_tim,   hide_index=True, use_container_width=True)
        with t2:
            st.dataframe(df_dosen, hide_index=True, use_container_width=True)

    # ── Build slot & preview kapasitas ──
    slot_dosen = build_slot_dosen(df_dosen, kolom_hari_lower, DURASI)
    render_preview_kapasitas(df_tim, slot_dosen, MAKS_PER_SESI)

    st.divider()

    # ── Proses ──
    if st.button(":rocket: Proses Penjadwalan", type="primary", use_container_width=True):
        with st.spinner("Memproses penjadwalan..."):
            slot_dosen_fresh = build_slot_dosen(df_dosen, kolom_hari_lower, DURASI)
            hasil = run_scheduling(df_tim, slot_dosen_fresh, MAKS_PER_SESI, RANDOM_SEED)

        if not hasil:
            st.error("Tidak ada tim yang berhasil dijadwalkan. Periksa kesesuaian data.")
            return

        df_hasil = pd.DataFrame(hasil)

        # ── Ringkasan ──
        _header_with_icon("chart-bar", "Ringkasan Hasil")
        total_tim     = len(df_tim)
        total_terplot = len(df_hasil)
        total_tidak   = total_tim - total_terplot

        m1, m2, m3 = st.columns(3)
        m1.metric("Total Tim", total_tim)
        m2.metric("Berhasil Terplot", total_terplot)
        m3.metric(
            "Tidak Terplot",
            total_tidak,
            delta=f"-{total_tidak}" if total_tidak > 0 else "0",
            delta_color="inverse" if total_tidak > 0 else "off",
        )

        with st.expander(":clipboard: Rekap per Hari & Bidang", expanded=False):
            pivot = pd.pivot_table(
                df_hasil, index="Hari", columns="Bidang",
                values="Ketua", aggfunc="count", fill_value=0,
            )
            st.dataframe(pivot, use_container_width=True)

        # ── Download ──
        st.divider()
        try:
            excel_bytes = export_excel(
                hasil, slot_dosen_fresh, df_tim, MAKS_PER_SESI, MAX_TABLE_ROW
            )
            st.download_button(
                label=":arrow_down: Download Jadwal Excel",
                data=excel_bytes,
                file_name=OUTPUT_FILE,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
            st.success(
                f"Penjadwalan selesai. **{total_terplot}** dari **{total_tim}** tim berhasil dijadwalkan."
            )
        except Exception as e:
            st.error(f"Gagal membuat file Excel: {e}")
            st.exception(e)


if __name__ == "__main__":
    main()
